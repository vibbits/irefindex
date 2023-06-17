# The GOEA is executed in a conda environment 
# Source: https://notebook.community/tanghaibao/goatools/notebooks/goea_nbt3102
# install goatools in the conda environment: conda install -c bioconda goatools
# to be able to use pickle in a conda environment, pickle5 should be installed
# conda install -c conda-forge pickle5
# First execute the conversion script to convert UniProt ids into gene ids. (UniProtID2GeneID.py)
from __future__ import print_function
import pickle5 as pickle 
from itertools import chain
import requests
import sys
from openpyxl import Workbook

path = '/home/guest/ETN-Silke/Scripts/UniProtKB2GeneID.pickle'

# Extract list of genes to use
genes = []
uniprot = []

# Load the pickle file to use
geneids = pickle.load(open(path,"rb"))
print(len(geneids))
# Make a distinction between key-value pairs of succesful and failed conversions of UniProtID to GeneID
result_key, result_values = next(iter(geneids.items()))
failed_key, failed_values = list(geneids.items())[1]
print(len(result_values))
print(len(failed_values))

# Iterate over the result values and make a list with the Gene ids and a list with the UniProt ids
for dict in result_values:
	genes.append(dict['to'])
	uniprot.append(dict['from'])
print(len(uniprot))
print(len(genes))

# Write UniProt ids and Gene ids to an Excel file
wb = Workbook()
ws = wb.active
for i in range(len(uniprot)):
	ws.cell(row=i+1, column=1, value=uniprot[i])
	ws.cell(row=i+1, column=2, value=genes[i])
wb.save("HumanGenes_imex.xlsx")

# Make background gene set
# Go to: https://www.ncbi.nlm.nih.gov/gene
# Search for: "9606"[Taxonomy ID] AND alive[property] AND genetype protein coding[Properties]
# Send the output of this search to a text file and save it (gene_result.txt)

# Move the background gene set to the current directory
# Command line: mv /home/guest/Downloads/gene_result.txt .

# Execute the following line to create a script genes_ncbi_homo_sapiens_proteincoding.py:
# python /home/guest/miniconda3/envs/GOanalysis/bin/ncbi_gene_results_to_python.py -o genes_ncbi_homo_sapiens_proteincoding.py gene_result.txt


# Download ontologies
from goatools.base import download_go_basic_obo
obo_fname = download_go_basic_obo()

# Download associations
url = 'https://ftp.ncbi.nlm.nih.gov/gene/DATA/gene2go.gz'
filename = url.split("/")[-1]
with open(filename, "wb") as f:
	r = requests.get(url)
	f.write(r.content)
fin_gene2go = open("gene2go",'r')

# Load ontologies 
from goatools.obo_parser import GODag 
obodag = GODag("go-basic.obo")

# Load associations
from goatools.anno.genetogo_reader import Gene2GoReader

# Read NCBI's gene2go. Store annotations in a list of named tuples
objanno = Gene2GoReader(fin_gene2go.name, taxids=[9606])

# Get namespace2association where:
#    namespace (ns) is:
#        BP: biological_process               
#        MF: molecular_function
#        CC: cellular_component
#    assocation (assc) is a dict:
#        key: NCBI GeneID
#        value: A set of GO IDs associated with that gene
ns2assoc = objanno.get_ns2assc()

# Loop over namespace and annotated genes
for nspc, id2gos in ns2assoc.items():
    print("{NS} {N:,} annotated human genes".format(NS=nspc, N=len(id2gos)))

# Load background gene set
from genes_ncbi_homo_sapiens_proteincoding import GENEID2NT as GeneID2nt_hs
print(len(GeneID2nt_hs))

# Initialize a GOEA (Gene Ontology Enrichment Analysis) object 
from goatools.goea.go_enrichment_ns import GOEnrichmentStudyNS
goeaobj = GOEnrichmentStudyNS(
	GeneID2nt_hs.keys(), # List of human protein-coding genes
	ns2assoc, # geneid/GO associations
	obodag, # Ontologies
	propagate_counts = False,
	alpha = 0.05, # default significance cut-off
	methods = ['fdr_bh'] # default multipletest correction method
)

# Read study genes
import os 
from openpyxl import load_workbook

geneid2uniprot = {}

# Get xlsx filename where data is stored
din_xlsx = "/home/guest/ETN-Silke/Scripts/HumanGenes_imex.xlsx"

# Read data and store gene ids and uniprot ids in a dictionary
if os.path.isfile(din_xlsx):
	wb = load_workbook(filename=din_xlsx)
	ws = wb.active
	for row in ws.iter_rows(min_row=2):
		uniprot,geneid = [cell.value for cell in row]
		if geneid:
			geneid2uniprot[int(geneid)] = uniprot
	print('{N} genes READ: {XLSX}'.format(N=len(geneid2uniprot),XLSX=din_xlsx))
else:
	raise RuntimeError('FILE NOT FOUND: {XLSX}'.format(XLSX=din_xlsx))

# Run Gene Ontology Enrichment Analysis (GOEA)
geneids_study = geneid2uniprot.keys()
goea_results_all = goeaobj.run_study(geneids_study)
goea_results_sig = [r for r in goea_results_all if r.p_fdr_bh < 0.05]

# Save the results in an xlsx file
goeaobj.wr_xlsx("GOEA_HumanGenes_imex.xlsx", goea_results_sig)