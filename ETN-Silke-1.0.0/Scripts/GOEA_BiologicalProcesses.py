# Only extract the biological processes from the GOEA result
from openpyxl import Workbook, load_workbook
import re
import sys
import pickle5 as pickle

# Read in the file to use
wb2read = load_workbook(filename="GOEA_HumanGenes_imex.xlsx")
wb2read_ws1 = wb2read['Sheet1']

# Activate the workbook and worksheet to use
wb_bp = Workbook()
ws_bp = wb_bp.active

# Add a header line to the file
header = [cell.value for cell in wb2read_ws1[1]]
ws_bp.append(header)

# Check for the biological processes only and extract these to a new file
for row in wb2read_ws1.iter_rows(min_row=2):
	if row[1].value == "BP":
		row_value = [cell.value for cell in row]
		ws_bp.append(row_value)

wb_bp.save("GOEA_BP_imex.xlsx")



