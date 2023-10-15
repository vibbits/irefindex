# Convert the ratio of each GO to a percentage and sort by percentage
from openpyxl import Workbook, load_workbook
from fractions import Fraction
import sys

# Load the Excel file 
wb2read = load_workbook(filename="GOEA_BP_all.xlsx")
wb2read_ws1 = wb2read['Sheet']

# Insert a new header for the percentage columns
wb2read_ws1.insert_cols(6)
wb2read_ws1.insert_cols(8)
wb2read_ws1.cell(row=1, column=6, value='Percentage study')
wb2read_ws1.cell(row=1, column=8, value='Percentage pop')

# Extract the ratios and transform them into percentages
for row in wb2read_ws1.iter_rows(min_row=2):
	ratio_study = row[4].value
	ratio_pop = row[6].value
	row = list(row)
	
	numerator, denominator = ratio_study.split("/")
	fraction = Fraction(int(numerator), int(denominator))
	perc_study = float(fraction)*100
	row[5].value = perc_study # insert the study percentage as an extra column
	
	numerator, denominator = ratio_pop.split("/")
	fraction = Fraction(int(numerator), int(denominator))
	perc_pop = float(fraction)*100
	row[7].value = perc_pop # insert the pop percentage as an extra column

# Sort by percentage in descending order and by significance level in ascending order
data = list(wb2read_ws1.values)

sorted_data = sorted(data[1:], key=lambda x: (-float(x[5]),x[11]))

# Extract the sorted data and write it to a new file
for i, row in enumerate(sorted_data, start=2):
	for j, cell in enumerate(row, start=1):
		wb2read_ws1.cell(row=i, column=j, value=cell)

wb2read.save("GOEA_BP_all_sorted2.xlsx")