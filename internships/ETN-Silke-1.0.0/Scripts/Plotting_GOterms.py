from openpyxl import load_workbook
import matplotlib.pyplot as plt

if __name__ == "__main__":
	# Load the Excel file 
	wb2read = load_workbook(filename="GOEA_BP_all_sorted2.xlsx")
	wb2read_ws1 = wb2read['Sheet']

	topGO_terms_names = {}
	# Define the GO-terms that should be plotted
	topterms = ["GO:0006355","GO:0006915","GO:0007165","GO:0006468","GO:0051301","GO:0010628"]

	# Iterate over the rows in the file and filter out the needed information for those GO-terms
	for row in wb2read_ws1.iter_rows(min_row=2):
		GOterm = row[0].value
		GOname = row[3].value
		perc = round(row[5].value,2)
		for term in topterms:
			if term in GOterm:
				topGO_terms_names[GOname] = perc
	print(topGO_terms_names)

	# Plot this data in a pie chart
	fig,ax = plt.subplots()
	ax.pie(list(topGO_terms_names.values()), labels=list(topGO_terms_names.keys()),autopct='%1.2f%%')
	plt.show()
